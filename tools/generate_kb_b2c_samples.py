from __future__ import annotations

import json
import re
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    load_workbook = None  # type: ignore[assignment]


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "frontend" / "src" / "app" / "openapi-test" / "samples.generated.json"
WORKBOOK_PATTERN = "KB_B2C_OpenAPI*.xlsx"
XML_DIR = ROOT / "KB_B2C"
LEGACY_XML_DIR = ROOT / "trx_b2c"
B2C_INDEX_FILE = "B2C.txt"
JSON_SPEC_ZIP_CANDIDATES = [ROOT / "json.zip", Path.home() / "Desktop" / "json.zip"]
JSON_SPEC_DIR_CANDIDATES = [ROOT / "json", Path.home() / "Desktop" / "json"]
INVESTMENT_WORKBOOK_PATTERNS = ["*API*.xlsx", "*api*.xlsx"]

INPUT_TEXT = "\uc785\ub825"
OUTPUT_TEXT = "\ucd9c\ub825"

DATA_HEADER = {
    "udId": "UDID",
    "subChannel": "subChannel",
    "deviceModel": "Android",
    "deviceOs": "Android",
    "carrier": "KT",
    "connectionType": "..",
    "appName": "..",
    "appVersion": "..",
    "scrNo": "0000",
}

B2C_BODY_OMITTED_FIELDS = {"gnl_ac_no", "gds_no", "pwd", "rnmcno"}
B2C_SPEC_OMITTED_FIELDS = {"rnmcno"}
CUSTOMER_ACCOUNT_CATEGORY = "\uace0\uac1d\uacc4\uc88c"
TRADING_CATEGORY = "\ud2b8\ub808\uc774\ub529"
INVESTMENT_INFO_CATEGORY = "\ud22c\uc790\uc815\ubcf4"
OTHER_CATEGORY = "\uae30\ud0c0\ud56d\ubaa9"
B2C_BUSINESS_CATEGORY_BY_CODE = {
    **dict.fromkeys(
        [
            "SPQN3390",
            "SSQM0004",
            "SSQM0005",
            "SSQM0006",
            "SSQM0009",
            "SSQM1801",
            "SSQM2932",
            "SSQM2952",
            "SSQN2952",
            "SWQA2301",
            "SWQB2301",
            "SWQM2302",
            "SWQM2412",
            "SWQN2302",
            "SZQM6019",
            "SSQM2121",
            "SSQM2392",
            "SSQM2442",
            "SSQM2443",
            "SSQM5472",
            "SKQO3390",
            "SPQM2205",
            "SPQM2206",
            "SPQM2207",
        ],
        CUSTOMER_ACCOUNT_CATEGORY,
    ),
    **dict.fromkeys(
        [
            "SSQM1802",
            "SSAM1801",
            "SSAM1802",
            "SSAM1805",
            "SSAM1806",
            "SSQN5472",
            "SSAM5762",
            "SSAM5763",
            "SSAM5764",
            "SSQM0831",
            "SSQM0832",
            "SSQM0833",
            "SSQM0834",
            "SSQM2341",
            "SSQM5475",
            "SSQM5765",
            "SSAM0831",
            "SKQM2106",
            "SKQM3350",
            "SPQM5472",
            "SPQM2220",
            "SPQM2226",
            "SPQM3390",
            "SPQO2226",
            "SRQM3051",
            "SPQM2106",
            "SKAM2101",
            "SKAM2102",
            "SPQN5472",
            "SKAM2201",
            "SKAM2202",
            "SPAO2104",
            "SPAO2106",
            "SPQM2204",
            "SPQM1818",
            "SPQM2103",
            "SPQM5473",
            "SPQN5473",
            "SPQO2105",
        ],
        TRADING_CATEGORY,
    ),
    **dict.fromkeys(
        [
            "SPAM2508",
            "SIAM4983",
            "GSS10030",
            "GSS10040",
            "GSA10020",
            "GSC10060",
            "SIQM4900",
            "SZQM0771",
            "IVU10140",
            "IVU10070",
            "IVU10080",
            "IVM10050",
            "IVS11560",
            "IVU10430",
            "IVU10420",
            "IVU10450",
            "IVU10020",
            "IVS11430",
            "IVS10920",
            "IVU10280",
            "IVU10270",
            "IVU10210",
            "IVU10240",
            "IVS10910",
            "IVS11190",
            "IVU10550",
            "IVSA0070",
            "IVA60140",
            "IVM30010",
            "IVA60190",
            "IVA10370",
        ],
        INVESTMENT_INFO_CATEGORY,
    ),
}


def is_filler_field_name(value: str) -> bool:
    return re.fullmatch(r"filler\d*", (value or "").strip().lower()) is not None


def should_omit_b2c_body_field(value: str) -> bool:
    normalized = (value or "").strip().lower()
    return normalized in B2C_BODY_OMITTED_FIELDS or is_filler_field_name(normalized)


def should_omit_b2c_spec_field(value: str) -> bool:
    normalized = (value or "").strip().lower()
    return normalized in B2C_SPEC_OMITTED_FIELDS or is_filler_field_name(normalized)


def b2c_business_category(tr_code: str, source_category: str = "") -> str:
    normalized_code = normalize_tr_code(tr_code)
    if normalized_code in B2C_BUSINESS_CATEGORY_BY_CODE:
        return B2C_BUSINESS_CATEGORY_BY_CODE[normalized_code]

    normalized_source = clean(source_category)
    if "\uacc4\uc88c" in normalized_source or "\uc794\uace0" in normalized_source:
        return CUSTOMER_ACCOUNT_CATEGORY
    if "\uc8fc\ubb38" in normalized_source:
        return TRADING_CATEGORY
    if "\uc2dc\uc138" in normalized_source or "\ud22c\uc790\uc815\ubcf4" in normalized_source:
        return INVESTMENT_INFO_CATEGORY
    return OTHER_CATEGORY


def clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_enabled_flag(value: Any) -> bool:
    return clean(value).upper() in {"1", "TRUE", "Y", "YES"}


def compact(value: str) -> str:
    return re.sub(r"\s+", "", value)


def normalize_tr_code(raw_code: str, sheet_name: str = "") -> str:
    source = raw_code or sheet_name
    source = re.sub(r"\([^)]*\)", "", source)
    source = re.sub(r"[^0-9A-Za-z]", "", source)
    return source.upper()


def normalize_field_key(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z]", "", value or "").lower()


def decode_text_bytes(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def first_existing_path(candidates: list[Path]) -> Path | None:
    for path in candidates:
        if path.exists():
            return path
    return None


def find_investment_workbook() -> Path | None:
    candidates: list[Path] = []
    for base_dir in (ROOT, Path.home() / "Desktop"):
        for pattern in INVESTMENT_WORKBOOK_PATTERNS:
            candidates.extend(path for path in base_dir.glob(pattern) if not path.name.startswith("~$"))
    if not candidates:
        return None

    for path in candidates:
        if "\uc2dc\uc138" in path.name:
            return path
    return sorted(candidates)[0]


def is_field_name(value: str) -> bool:
    if not value:
        return False
    upper = value.upper()
    if upper in {"\uc601\ubb38\uba85", "INPUT", "OUTPUT", "IN", "OUT", "RQ", "RP"}:
        return False
    return bool(re.search(r"[A-Za-z]", value))


def field_from_standard_row(row: tuple[Any, ...]) -> dict[str, str] | None:
    korean = clean(row[2] if len(row) > 2 else "")
    english = clean(row[3] if len(row) > 3 else "")
    if not is_field_name(english):
        return None
    return {
        "korean": korean,
        "name": english,
        "type": clean(row[4] if len(row) > 4 else ""),
        "length": clean(row[5] if len(row) > 5 else ""),
        "decimal": clean(row[6] if len(row) > 6 else ""),
        "note": clean(row[7] if len(row) > 7 else ""),
        "description": clean(row[7] if len(row) > 7 else ""),
        "required": "N",
    }


def field_from_realtime_row(row: tuple[Any, ...]) -> dict[str, str] | None:
    korean = clean(row[1] if len(row) > 1 else "")
    english = clean(row[6] if len(row) > 6 else "")
    if not is_field_name(english):
        return None
    return {
        "korean": korean,
        "name": english,
        "type": "",
        "length": "",
        "decimal": "",
        "note": "",
        "description": "",
        "required": "N",
    }


def has_marker(row: tuple[Any, ...], marker: str) -> bool:
    return any(clean(cell).upper() == marker for cell in row)


def has_text(row: tuple[Any, ...], text: str) -> bool:
    return any(text in clean(cell) for cell in row)


def parse_block_fields(rows: list[tuple[Any, ...]]) -> list[dict[str, str]]:
    fields: list[dict[str, str]] = []

    for index, row in enumerate(rows):
        if not has_marker(row, "BLOCK") or not has_text(row, INPUT_TEXT):
            continue

        for field_row in rows[index + 2 :]:
            if has_marker(field_row, "BLOCK"):
                break
            if not any(clean(cell) for cell in field_row):
                break
            field = field_from_standard_row(field_row)
            if field:
                fields.append(field)

    return fields


def parse_rq_fields(rows: list[tuple[Any, ...]]) -> list[dict[str, str]]:
    fields: list[dict[str, str]] = []
    collecting = False
    started_with = ""

    for row in rows:
        marker = clean(row[1] if len(row) > 1 else "").upper()
        if marker in {"RQ", "IN"}:
            if fields:
                break
            collecting = True
            started_with = marker
            continue

        if collecting and marker in {"RP", "OUT"}:
            break
        if collecting and has_text(row, OUTPUT_TEXT):
            break

        if not collecting:
            continue

        if not any(clean(cell) for cell in row):
            if fields:
                break
            continue

        field = field_from_standard_row(row)
        if field:
            fields.append(field)

    return fields if started_with else []


def parse_realtime_key_fields(rows: list[tuple[Any, ...]]) -> list[dict[str, str]]:
    fields: list[dict[str, str]] = []
    collecting = False

    for row in rows:
        row_text = " ".join(clean(cell) for cell in row)
        if INPUT_TEXT in row_text and "input" in row_text.lower():
            collecting = True
            continue
        if collecting and (OUTPUT_TEXT in row_text or "output" in row_text.lower()):
            break
        if not collecting:
            continue

        field = field_from_realtime_row(row)
        if field:
            fields.append(field)

    return fields


def parse_input_fields(ws) -> tuple[str, list[dict[str, str]]]:
    rows = list(ws.iter_rows(values_only=True))

    block_fields = parse_block_fields(rows)
    if block_fields:
        return "block", block_fields

    rq_fields = parse_rq_fields(rows)
    if rq_fields:
        return "rq", rq_fields

    realtime_fields = parse_realtime_key_fields(rows)
    if realtime_fields:
        return "realtime", realtime_fields

    return "unknown", []


def extract_description(ws, fallback: str) -> str:
    rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
    for row in rows:
        cells = [clean(cell) for cell in row]
        for index, cell in enumerate(cells):
            if cell == "\uc124\uba85":
                for value in cells[index + 1 :]:
                    if value:
                        return value
    return fallback


def default_value(field: dict[str, str]) -> str:
    name = field["name"]
    upper = name.upper()
    lower = name.lower()
    korean = field.get("korean", "")
    note = field.get("note", "")
    configured_default = clean(field.get("default", ""))
    text = f"{upper} {korean} {note}"
    compact_text = compact(text)

    if configured_default:
        return configured_default
    if upper.startswith("ATTR_"):
        return ""
    if upper in {"FILLER", "CONTKEY"} or "CONTKEY" in upper:
        return ""
    if upper in {"GNL_AC_NO", "AC_NO", "ACNO"} or upper.endswith("_AC_NO"):
        return "{{gnlAcNo}}"
    if upper == "GDS_NO":
        return "{{gdsNo}}"
    if upper in {"PWD", "AC_PWD"} or upper.endswith("_PWD"):
        return "{{pwd}}"
    if upper in {"CI_NO", "CINO"}:
        return "{{ciNo}}"
    if upper == "CONTIF":
        return "0"
    if upper in {"RL_TM_KEY", "TR_KEY"}:
        if "\ud574\uc678" in text or "KRX_CD" in text or "\uac70\ub798\uc18c" in text:
            return "NASAAPL"
        return "005930"
    if upper in {"IS_CD", "PDNO", "STK_CD", "STCK_SHRN_ISCD"} or "IS_CD" in upper:
        return "005930"
    if upper in {"KRX_CD", "EXCH_CD"}:
        return "NAS"
    if upper == "NTN_CD":
        return "USA"
    if upper == "N_MINUTE":
        return "1"
    if upper in {"INQ_CNT", "RCRD_C"} or upper.endswith("_CNT") or "COUNT" in upper:
        return "10"
    if "STRT_DT" in upper or "START_DT" in upper or "\uc2dc\uc791\uc77c\uc790" in compact_text:
        return "20250101"
    if "END_DT" in upper or "\uc885\ub8cc\uc77c\uc790" in compact_text:
        return "20251231"
    if upper in {"DT", "BSNSS_DT", "KOR_DT"} or upper.endswith("_DT"):
        return "20250617"
    if upper in {"TM", "KOR_TM"} or upper.endswith("_TM"):
        return "090000"
    if upper.endswith("_YN"):
        return "N"
    if any(token in upper for token in ("CLSF", "CCD", "DVSN", "GB", "TYPE")):
        return "1"
    if upper.endswith("_Q") or "QTY" in upper:
        return "1"
    if upper.endswith("_C") or "CNT" in upper:
        return "10"
    if upper.startswith("N_") and lower != "n_minute":
        return "1"

    return ""


def build_data_body(fields: list[dict[str, str]]) -> dict[str, str]:
    body: dict[str, str] = {}
    for field in fields:
        key = field["name"]
        upper = key.upper()
        if (
            is_enabled_flag(field.get("skipValue"))
            or should_omit_b2c_body_field(key)
            or upper == "PWD"
            or upper.endswith("_PWD")
        ):
            continue
        if key not in body:
            body[key] = default_value(field)
    return body


def load_b2c_catalog(workbook_path: Path) -> dict[str, Any]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to load the KB B2C workbook.")

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    index = wb.worksheets[0]
    b2c: list[dict[str, Any]] = []
    layout_counts: dict[str, int] = {}

    for row in index.iter_rows(min_row=3, values_only=True):
        raw_code = clean(row[1] if len(row) > 1 else "")
        service_name = clean(row[2] if len(row) > 2 else "")
        sheet_name = clean(row[3] if len(row) > 3 else "") or raw_code
        if not raw_code:
            continue
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        tr_code = normalize_tr_code(raw_code, sheet_name)
        layout, fields = parse_input_fields(ws)
        description = extract_description(ws, service_name)
        layout_counts[layout] = layout_counts.get(layout, 0) + 1

        body = {
            "dataHeader": dict(DATA_HEADER),
            "dataBody": build_data_body(fields),
        }

        label_prefix = "\uc2e4\uc2dc\uac04 " if layout in {"realtime", "unknown"} and tr_code.startswith("KBRS") else ""
        b2c.append(
            {
                "id": f"Tkb_{tr_code}_B2C",
                "label": f"{tr_code} {label_prefix}{service_name}".strip(),
                "method": "POST",
                "endpoint": f"/excel-b2c/{tr_code.lower()}",
                "description": f"B2C \ud22c\uc790\uc815\ubcf4: {description}",
                "headers": {"Content-Type": "application/json"},
                "body": body,
                "query": {},
                "fileName": workbook_path.name,
                "source": "kb-b2c-excel",
                "layout": layout,
                "sheetName": sheet_name,
            }
        )

    return {"b2c": b2c, "layoutCounts": layout_counts}


def read_xml_text(path: Path) -> tuple[str, str]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace"), "utf-8-replace"


def read_text_file(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


FieldMetadata = dict[str, str]
FieldDescriptionLookup = dict[str, dict[str, dict[str, FieldMetadata]]]


def required_text(value: Any) -> str:
    text = clean(value)
    upper = text.upper()
    if not text:
        return ""
    if upper in {"1", "Y", "YES", "TRUE", "M", "MANDATORY", "REQUIRED"} or "\ud544\uc218" in text:
        return "Y"
    return "N"


def merge_description(
    descriptions: FieldDescriptionLookup,
    tr_code: str,
    section: str,
    field_name: str,
    description: str,
    required: str = "",
    overwrite_description: bool = True,
) -> None:
    normalized_code = normalize_tr_code(tr_code)
    normalized_field = normalize_field_key(field_name)
    text = clean(description)
    required_flag = required_text(required)
    if not normalized_code or not normalized_field or (not text and not required_flag):
        return
    field_metadata = descriptions.setdefault(normalized_code, {"input": {}, "output": {}}).setdefault(section, {}).setdefault(
        normalized_field, {}
    )
    if text and (overwrite_description or not clean(field_metadata.get("description"))):
        field_metadata["description"] = text
    if required_flag:
        field_metadata["required"] = required_flag


def merge_json_item_descriptions(
    descriptions: FieldDescriptionLookup,
    item: dict[str, Any],
    fallback_code: str,
) -> None:
    tr_code = normalize_tr_code(clean(item.get("code")) or fallback_code)
    if not tr_code:
        return
    for section, map_key in (("input", "inputMap"), ("output", "outputMap")):
        field_map = item.get(map_key)
        if not isinstance(field_map, dict):
            continue
        for field_name, field_spec in field_map.items():
            if not isinstance(field_spec, dict):
                continue
            merge_description(
                descriptions,
                tr_code,
                section,
                str(field_name),
                clean(field_spec.get("etc")),
                clean(field_spec.get("need")),
            )


def load_json_directory_descriptions(path: Path) -> FieldDescriptionLookup:
    descriptions: FieldDescriptionLookup = {}
    for json_path in sorted(path.rglob("*.json")):
        try:
            item = json.loads(decode_text_bytes(json_path.read_bytes()))
        except (json.JSONDecodeError, UnicodeDecodeError):
            continue
        if not isinstance(item, dict):
            continue
        merge_json_item_descriptions(descriptions, item, json_path.stem)
    return descriptions


def load_json_zip_descriptions(path: Path) -> FieldDescriptionLookup:
    descriptions: FieldDescriptionLookup = {}
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            if not name.lower().endswith(".json"):
                continue
            try:
                item = json.loads(decode_text_bytes(archive.read(name)))
            except (json.JSONDecodeError, UnicodeDecodeError):
                continue
            if not isinstance(item, dict):
                continue

            merge_json_item_descriptions(descriptions, item, Path(name).stem)
    return descriptions


def field_name_and_description_from_workbook_row(row: tuple[Any, ...]) -> tuple[str, str] | None:
    cells = [clean(cell) for cell in row]
    if len(cells) > 3 and is_field_name(cells[3]):
        return cells[3], clean(cells[7] if len(cells) > 7 else "")
    if len(cells) > 6 and is_field_name(cells[6]):
        return cells[6], clean(cells[12] if len(cells) > 12 else cells[7] if len(cells) > 7 else "")
    return None


def workbook_section_from_row(row: tuple[Any, ...], current_section: str) -> str:
    cells = [clean(cell) for cell in row]
    row_text = " ".join(cells)
    upper_cells = {cell.upper() for cell in cells}
    if (
        INPUT_TEXT in row_text
        or "RQ" in upper_cells
        or "IN" in upper_cells
        or any(cell.lower() == "input" for cell in cells)
    ):
        return "input"
    if (
        OUTPUT_TEXT in row_text
        or "RP" in upper_cells
        or "OUT" in upper_cells
        or any(cell.lower() == "output" for cell in cells)
    ):
        return "output"
    return current_section


def load_investment_workbook_descriptions(path: Path) -> FieldDescriptionLookup:
    descriptions: FieldDescriptionLookup = {}
    if load_workbook is None:
        return descriptions

    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet in workbook.worksheets:
        tr_code = normalize_tr_code(sheet.title)
        if not tr_code:
            continue
        section = ""
        for row in sheet.iter_rows(values_only=True):
            section = workbook_section_from_row(row, section)
            if not section:
                continue
            parsed = field_name_and_description_from_workbook_row(row)
            if not parsed:
                continue
            field_name, description = parsed
            merge_description(descriptions, tr_code, section, field_name, description)
    return descriptions


def load_field_descriptions() -> tuple[FieldDescriptionLookup, dict[str, str]]:
    descriptions: FieldDescriptionLookup = {}
    sources: dict[str, str] = {}

    json_dir = first_existing_path(JSON_SPEC_DIR_CANDIDATES)
    if json_dir:
        descriptions = load_json_directory_descriptions(json_dir)
        sources["jsonDirectory"] = str(json_dir)
    else:
        json_zip = first_existing_path(JSON_SPEC_ZIP_CANDIDATES)
        if json_zip:
            descriptions = load_json_zip_descriptions(json_zip)
            sources["jsonZip"] = str(json_zip)

    investment_workbook = find_investment_workbook()
    if investment_workbook:
        investment_descriptions = load_investment_workbook_descriptions(investment_workbook)
        for tr_code, sections in investment_descriptions.items():
            for section, fields in sections.items():
                for field_name, metadata in fields.items():
                    merge_description(
                        descriptions,
                        tr_code,
                        section,
                        field_name,
                        clean(metadata.get("description")),
                        overwrite_description=False,
                    )
        sources["investmentWorkbook"] = str(investment_workbook)

    return descriptions, sources


def apply_field_descriptions(tr_code: str, section: str, fields: list[dict[str, str]], descriptions: FieldDescriptionLookup) -> None:
    field_metadata = descriptions.get(normalize_tr_code(tr_code), {}).get(section, {})
    for field in fields:
        metadata = field_metadata.get(normalize_field_key(field.get("name", "")), {})
        field["description"] = clean(metadata.get("description")) or clean(field.get("description"))
        field["required"] = required_text(metadata.get("required")) or field.get("required") or "N"


def count_field_descriptions(items: list[dict[str, Any]]) -> int:
    count = 0
    for item in items:
        for section_name in ("inputSpec", "outputSpec"):
            for field in item.get(section_name, []):
                if clean(field.get("description")):
                    count += 1
    return count


def load_b2c_index(xml_dir: Path) -> list[dict[str, Any]]:
    index_path = xml_dir / B2C_INDEX_FILE
    if not index_path.exists():
        return []

    entries: list[dict[str, Any]] = []
    current_category = ""
    duplicate_counts: dict[str, int] = {}

    for line in read_text_file(index_path).splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith("[") and stripped.endswith("]"):
            current_category = stripped.strip("[]").strip()
            continue
        if "-" not in stripped:
            continue

        raw_code, raw_label = stripped.split("-", 1)
        tr_code = normalize_tr_code(raw_code)
        if not tr_code:
            continue

        duplicate_counts[tr_code] = duplicate_counts.get(tr_code, 0) + 1
        entries.append(
            {
                "trCode": tr_code,
                "label": clean(raw_label) or tr_code,
                "category": current_category,
                "order": len(entries),
                "duplicateIndex": duplicate_counts[tr_code],
            }
        )

    return entries


def parse_xml_root(path: Path) -> tuple[ET.Element, str]:
    text, encoding = read_xml_text(path)
    xml_body = re.sub(r"<\?xml[^>]*\?>", "", text, count=1).lstrip()
    return ET.fromstring(xml_body), encoding


def field_from_xml_node(node: ET.Element) -> dict[str, str] | None:
    name = clean(node.attrib.get("Name"))
    if not name or name.startswith("_"):
        return None
    if should_omit_b2c_spec_field(name):
        return None
    if is_enabled_flag(node.attrib.get("SkipValue")):
        return None

    io_type = clean(node.attrib.get("IOType")).lower()
    spec = clean(node.attrib.get("Spec") or node.attrib.get("format"))
    if io_type == "record" or spec.lower() == "struct":
        return None

    return {
        "korean": clean(node.attrib.get("Desc")),
        "name": name,
        "type": clean(node.attrib.get("format") or node.attrib.get("Spec")),
        "length": clean(node.attrib.get("Size") or node.attrib.get("SizeEN") or node.attrib.get("SizeKR")),
        "decimal": clean(node.attrib.get("portion")),
        "note": clean(node.attrib.get("IOType")),
        "default": clean(node.attrib.get("Default")),
        "skipValue": clean(node.attrib.get("SkipValue")),
        "description": "",
        "required": "N",
    }


def parse_xml_fields(root: ET.Element, section_name: str) -> list[dict[str, str]]:
    section_node = root.find(section_name)
    if section_node is None:
        return []

    fields: list[dict[str, str]] = []
    for node in section_node.iter():
        if node is section_node:
            continue
        field = field_from_xml_node(node)
        if field:
            fields.append(field)
    return fields


def build_b2c_xml_item(
    xml_path: Path,
    root: ET.Element,
    encoding: str,
    metadata: dict[str, Any] | None = None,
    descriptions: FieldDescriptionLookup | None = None,
) -> dict[str, Any]:
    metadata = metadata or {}
    tr_code = normalize_tr_code(
        clean(root.findtext("Name")) or clean(root.findtext("Route")) or clean(metadata.get("trCode")) or xml_path.stem
    )
    description = clean(metadata.get("label")) or clean(root.findtext("Desc")) or tr_code
    category = clean(metadata.get("category"))
    duplicate_index = int(metadata.get("duplicateIndex") or 1)
    input_fields = parse_xml_fields(root, "Input")
    output_fields = parse_xml_fields(root, "Output")
    descriptions = descriptions or {}
    apply_field_descriptions(tr_code, "input", input_fields, descriptions)
    apply_field_descriptions(tr_code, "output", output_fields, descriptions)
    body = {
        "dataHeader": dict(DATA_HEADER),
        "dataBody": build_data_body(input_fields),
    }
    item_id = xml_path.stem if duplicate_index <= 1 else f"{xml_path.stem}__{duplicate_index}"
    description_prefix = f"{category}: " if category else ""

    return {
        "id": item_id,
        "label": f"{tr_code} {description}".strip(),
        "method": "POST",
        "endpoint": f"/api/v1/{tr_code.lower()}",
        "transactionCode": tr_code,
        "businessCategory": b2c_business_category(tr_code, category),
        "description": f"B2C XML 전문: {description_prefix}{description}",
        "headers": {"Content-Type": "application/json"},
        "body": body,
        "query": {},
        "inputSpec": input_fields,
        "outputSpec": output_fields,
        "fileName": xml_path.name,
        "source": "kb-b2c-xml",
        "layout": "xml",
        "inputFieldCount": len(input_fields),
        "outputFieldCount": len(output_fields),
        "menu": category,
        "sourceEncoding": encoding,
    }


def load_b2c_xml_catalog(xml_dir: Path, descriptions: FieldDescriptionLookup | None = None) -> dict[str, Any]:
    b2c: list[dict[str, Any]] = []
    encodings: dict[str, int] = {}
    index_entries = load_b2c_index(xml_dir)
    xml_entries: dict[str, tuple[Path, ET.Element, str]] = {}

    for xml_path in sorted(xml_dir.glob("*.xml")):
        root, encoding = parse_xml_root(xml_path)
        encodings[encoding] = encodings.get(encoding, 0) + 1
        tr_code = normalize_tr_code(clean(root.findtext("Name")) or clean(root.findtext("Route")) or xml_path.stem)
        xml_entries[tr_code] = (xml_path, root, encoding)

    consumed_codes: set[str] = set()
    for entry in index_entries:
        tr_code = clean(entry.get("trCode"))
        if tr_code in consumed_codes:
            continue
        xml_entry = xml_entries.get(tr_code)
        if not xml_entry:
            continue
        xml_path, root, encoding = xml_entry
        b2c.append(build_b2c_xml_item(xml_path, root, encoding, entry, descriptions))
        consumed_codes.add(tr_code)

    for tr_code, (xml_path, root, encoding) in sorted(xml_entries.items()):
        if tr_code in consumed_codes:
            continue
        b2c.append(build_b2c_xml_item(xml_path, root, encoding, descriptions=descriptions))

    return {"b2c": b2c, "encodings": encodings, "indexCount": len(index_entries)}


def merge_b2c_catalogs(excel_items: list[dict[str, Any]], xml_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    excel_by_id = {clean(item.get("id")): item for item in excel_items}
    merged: list[dict[str, Any]] = []

    for xml_item in xml_items:
        existing = excel_by_id.get(clean(xml_item.get("id")))
        if existing:
            next_item = dict(existing)
            next_item["source"] = "kb-b2c-excel+xml"
            next_item["sourceXmlFile"] = xml_item.get("fileName")
            next_item["xmlInputFieldCount"] = xml_item.get("inputFieldCount")
            merged.append(next_item)
        else:
            merged.append(xml_item)

    return merged


def main() -> None:
    xml_dir = XML_DIR if XML_DIR.exists() else LEGACY_XML_DIR
    field_descriptions, description_sources = load_field_descriptions()
    xml_loaded = (
        load_b2c_xml_catalog(xml_dir, field_descriptions)
        if xml_dir.exists()
        else {"b2c": [], "encodings": {}, "indexCount": 0}
    )
    workbook_path: Path | None = None
    loaded: dict[str, Any] = {"b2c": [], "layoutCounts": {}}

    if xml_loaded["b2c"]:
        b2c_catalog = xml_loaded["b2c"]
    else:
        matches = sorted(ROOT.glob(WORKBOOK_PATTERN))
        if not matches:
            raise SystemExit(f"KB B2C source was not found: {XML_DIR} or {WORKBOOK_PATTERN}")
        workbook_path = matches[0]
        loaded = load_b2c_catalog(workbook_path)
        b2c_catalog = loaded["b2c"]

    catalog = {
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "sourceWorkbook": str(workbook_path) if workbook_path else "",
        "sourceXmlDirectory": str(xml_dir) if xml_loaded["b2c"] else "",
        "descriptionSources": description_sources,
        "b2c": b2c_catalog,
    }

    OUTPUT.write_text(json.dumps(catalog, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    print(f"B2C: {len(catalog['b2c'])}")
    if loaded["layoutCounts"]:
        print("Layouts:", loaded["layoutCounts"])
    if xml_loaded["b2c"]:
        print(f"XML samples: {len(xml_loaded['b2c'])}, index entries: {xml_loaded.get('indexCount', 0)}, encodings: {xml_loaded['encodings']}")
    if description_sources:
        print(f"Description fields: {count_field_descriptions(catalog['b2c'])}, sources: {description_sources}")


if __name__ == "__main__":
    main()
